from openpyxl import load_workbook
import xml.etree.ElementTree as ET

# Configurable parts
XML_file_name = 'ExpenseManagerConfigFile.xml'

# Classes:
class Expense:
    def __init__(self, BusinessName, TransactionSum, ChargeAmount) :
        self.BusinessName = BusinessName
        self.TransactionSum = TransactionSum
        self.ChargeAmount = ChargeAmount

class CategoryCounter:
    def __init__(self, CategoryName, Counter) :
        self.CategoryName = CategoryName
        self.Counter = Counter
    
    def __lt__(self, other):
        return self.CategoryName < other.CategoryName

# Helpers functions:
def is_this_string_contains_hebrew_char(s):
    return any("\u0590" <= c <= "\u05EA" for c in s)
    
def create_expenses_list_from_excel():
    ThisMonthExpensesFileName = "ThisMonthExpenses.xlsx"
    wb = load_workbook(filename = ThisMonthExpensesFileName)
    ws = wb.active
    current_month_expenses_list = []
    for row in ws.iter_rows(min_row=4, max_col=5, max_row=10):
        if is_this_string_contains_hebrew_char(str(row[1].value)):
            BusinessName = str(row[1].value)[::-1]
        else:
            BusinessName = str(row[1].value)
        TransactionSum = str(row[2].value)
        ChargeAmount = str(row[3].value)
        current_month_expenses_list.append(Expense(BusinessName, TransactionSum, ChargeAmount))
    
    return current_month_expenses_list

def initiallize_counters_list_for_known_categories():
    XML_tree = ET.parse(XML_file_name)
    XML_root = XML_tree.getroot()
    categories_counters_list = []

    for category in XML_root.iter('Category'):
        should_i_insert = True
        for cc in categories_counters_list:
            if cc.CategoryName == category.text:
                should_i_insert = False
                break
        if should_i_insert:
            categories_counters_list.append(CategoryCounter(category.text, 0))
    categories_counters_list.sort()
    return categories_counters_list

def get_the_business_category_if_it_known(XML_root, business_name):
    is_the_business_of_this_expense_known = False
    business_category = ""
    for business in XML_root:
            if business[0].text == business_name:
                is_the_business_of_this_expense_known = True
                business_category = business[1].text
                break
    return is_the_business_of_this_expense_known, business_category

def sum_expense(charge_amount, category, categories_counters_list):
    is_it_succeed = False
    for category_counter in categories_counters_list:
        if category_counter.CategoryName == category:
            category_counter.Counter = float(category_counter.Counter) + float(charge_amount)
            is_it_succeed = True
            break
    return is_it_succeed

def ask_user_to_which_category_belongs_the_expense(categories_counters_list, expense):
    is_it_new_category = False
    print("\nhello, from the following categories: ")
    i = 0
    for categotyCounter in categories_counters_list:
        print(str(i) + ". " +categotyCounter.CategoryName)
        i= i+1
    print(str(i) + ". " + "New Category\n")
    try:
        categoryNumber = input("to which one belongs the business named: " +  expense.BusinessName + "?\n")    
        categoryNumber = int(categoryNumber)
    except ValueError:
        categoryNumber = input("Insert number between 0 to "+str(i) + " please.\n")
        categoryNumber = int(categoryNumber)
    if categoryNumber<0 or categoryNumber>i :
        categoryNumber = input("illegal number")
    
    # If we arrive here The user gave us valid numebr
    if categoryNumber < i:
        category =  categories_counters_list[categoryNumber].CategoryName
    else:
        category = input("so what is the category of this business then ?\n")
        print("\n")
        is_it_new_category = True

    return category, is_it_new_category
    
def print_summary_of_expenses(categories_counters_list):
    print("your total is: ")
    for categotyCounter in categories_counters_list:
        print("wasted on " + categotyCounter.CategoryName +" " + str(categotyCounter.Counter) +  "NIS")

def save_the_category_of_the_business_in_the_xml(XML_tree):
    new_business = ET.SubElement(XML_tree.getroot(), 'Business')
    ET.SubElement(new_business, 'BusinessName').text = expense.BusinessName
    ET.SubElement(new_business, 'Category').text = category

def save_all_xml_changes(XML_tree):
        XML_tree.write("ExpenseManagerConfigFile.xml")

if __name__ == "__main__":
    XML_tree = ET.parse(XML_file_name) # Get the configuration XML file as workable tree:
    current_month_expenses_list = create_expenses_list_from_excel()
    categories_counters_list = initiallize_counters_list_for_known_categories()

    for expense in current_month_expenses_list:
        is_the_business_of_this_expense_known, category = get_the_business_category_if_it_known(XML_tree.getroot(), expense.BusinessName)
        if is_the_business_of_this_expense_known == False:
            category, is_it_new_category = ask_user_to_which_category_belongs_the_expense(categories_counters_list, expense)
            if is_it_new_category:
                categories_counters_list.append(CategoryCounter(category,0))
                categories_counters_list.sort()
            save_the_category_of_the_business_in_the_xml(XML_tree)      
        sum_expense(expense.ChargeAmount, category, categories_counters_list)

    print_summary_of_expenses(categories_counters_list)
    save_all_xml_changes(XML_tree)